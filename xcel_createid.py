# -*- coding: utf-8 -*-
"""Xcel_CreateId.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1czUOOoDrpWPL18Ti07aY8loJGzN00t8c

使用前に  
・dateカラムとidカラムは「セルの書式設定」で「文字列」にする（議案データであればB, D, N列）  
・dateカラムが４桁かつ月日がバラバラになっていないか確認  
　　例）123は1月23日の「0123」なのか12月3日の「1203」なのか分からないため  
・編集したいファイルを予めGoogleドライブに保存

１．はじめに Google ドライブのファイルへのアクセスを許可する（左のドライブファイルからGoogle ドライブをマウントする）
"""

!pip install openpyxl

"""２．編集したいファイルパスを指定（左のフォルダから編集したいファイルを副ボタンでクリック。「パスをコピー」し、' 'の中に貼り付け）  
３．保存先のファイルパスと好きなファイル名を指定
"""

import openpyxl

import_path = r'/content/drive/My Drive/iinkai_test.xlsx' # 編集したいファイルのパスを作成
wb = openpyxl.load_workbook(import_path, data_only=True)
ws = wb['Sheet1'] # 編集したいシート名
export_path = r'/content/drive/My Drive/iinkai_result.xlsx' # 保存先のファイルパスとファイル名を指定

"""４．IDを作成したいデータのプログラムを実行（議案、委員会）

議案データ
"""

def is_empty(cell):
  return cell.value is None or not str(cell.value).strip()

i = 2
max = ws.max_column
for row in ws.iter_rows(min_row=2):
    # 空行になったら終了
    if all(is_empty(c) for c in row):
      break
    # 下３桁の数字リセットサイン
    date_cell = str(ws.cell(row=i, column=2).value)
    pre_date_cell = str(ws.cell(row=i-1, column=2).value)
    if date_cell == pre_date_cell:
      ws.cell(row=i, column=max+1).value = 1
    else:
      ws.cell(row=i, column=max+1).value = 0
    # 下３桁の数字作成
    if ws.cell(row=i, column=max+1).value == 0:
      ws.cell(row=i, column=max+2).value = 1
    else:
      ws.cell(row=i, column=max+2).value = ws.cell(row=i-1, column=max+2).value + 1
    # id作成
    year = str(ws.cell(row=i, column=1).value)
    date = str(ws.cell(row=i, column=2).value).zfill(4)
    jis = str(ws.cell(row=i, column=3).value)
    num = str(ws.cell(row=i, column=max+2).value).zfill(3)
    id = year + date + jis + num
    ws.cell(row=i, column=4).value = str(id)
    ws.cell(row=i, column=max+3).value = str(id)
    i+=1

    values = []
    for col in row:
      values.append(col.value)
    print(values)
    wb.save(export_path)

"""委員会データ

４-1．ID作成したい市町村コードを下記より指定  

自治体コード	団体名
  
滋賀県  
25201	大津市　25202	 彦根市　25203	長浜市　25204	近江八幡市	  
25206	草津市　25207	守山市　25208	栗東市　25209	甲賀市	  
25210	野洲市　25211	湖南市　25212	高島市　25213	東近江市	  
25214	米原市	  

奈良県  
29201	奈良市　29202	大和高田市　29203	大和郡山市　29204	天理市	  
29205	橿原市　29206	桜井市　29207	五條市　29208	御所市  
29209	生駒市　29210	香芝市　29211	葛城市　29212	宇陀市	  

和歌山県  
30201	和歌山市　30202	海南市　30203	橋本市　30204	有田市	  
30205	御坊市　30206	田辺市　30207	新宮市　30208	紀の川市	  
30209	岩出市
"""

code = 30202 #市町村コードの指定

def is_empty(cell):
  return cell.value is None or not str(cell.value).strip()

i = 2
max = ws.max_column
for row in ws.iter_rows(min_row=2):
    # 空行になったら終了
    if all(is_empty(c) for c in row):
      break
    # 下３桁の数字リセットサイン
    date_cell = str(ws.cell(row=i, column=2).value)
    pre_date_cell = str(ws.cell(row=i-1, column=2).value)
    if date_cell == pre_date_cell:
      ws.cell(row=i, column=max+1).value = 1
    else:
      ws.cell(row=i, column=max+1).value = 0
    # 下３桁の数字作成
    if ws.cell(row=i, column=max+1).value == 0:
      ws.cell(row=i, column=max+2).value = 1
    else:
      ws.cell(row=i, column=max+2).value = ws.cell(row=i-1, column=max+2).value + 1
    # id作成
    date = str(ws.cell(row=i, column=2).value)
    jis = str(code)
    num = str(ws.cell(row=i, column=max+2).value).zfill(3)
    id = date + jis + num
    ws.cell(row=i, column=3).value = str(id)
    ws.cell(row=i, column=max+3).value = str(id)
    i+=1

    values = []
    for col in row:
      values.append(col.value)
    print(values)
    wb.save(export_path)